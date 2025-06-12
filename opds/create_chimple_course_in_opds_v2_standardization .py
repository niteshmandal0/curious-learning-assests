import os
import json
from datetime import datetime
from openpyxl import load_workbook
from urllib.parse import urljoin

# Configuration
EXCEL_FILE = 'Respect Course Latest All Course Details From dashboard.xlsx'
BASE_URL = 'http://chimple.cc/opds/'  # Base URL for your OPDS catalog
TYPE_OPDS = 'application/opds+json'
PUB_TYPE = 'application/opds-publication+json'
SKIP_SHEETS = {'All Courses', 'Sheet4', 'Sheet5'}  # adjust as needed

# Output folders
OUTPUT_DIR = 'output_opds_combined'
GRADE_DIR = os.path.join(OUTPUT_DIR, 'grades')
LESSON_DIR = os.path.join(OUTPUT_DIR, 'lessons')
IMAGE_DIR = os.path.join(OUTPUT_DIR, 'images')

os.makedirs(GRADE_DIR, exist_ok=True)
os.makedirs(LESSON_DIR, exist_ok=True)
os.makedirs(IMAGE_DIR, exist_ok=True)

# Load workbook
print(f"Loading workbook: {EXCEL_FILE}")
wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)

# --- Generate index.json ---
navigation = []
for sheet_name in wb.sheetnames:
    if sheet_name in SKIP_SHEETS:
        continue
    filename = sheet_name.replace(' ', '').lower() + '.json'
    navigation.append({
        'href': filename,
        'title': sheet_name,
        'type': TYPE_OPDS
    })
index = {
    'metadata': {'title': 'Chimple Learning'},
    'links': [{'rel': 'self', 'href': urljoin(BASE_URL, 'index.json'), 'type': TYPE_OPDS}],
    'navigation': navigation
}
with open(os.path.join(OUTPUT_DIR, 'index.json'), 'w', encoding='utf-8') as f:
    json.dump(index, f, indent=2)
print(f"Generated index.json with {len(navigation)} grades.")

# --- Process sheets ---
for sheet_name in wb.sheetnames:
    if sheet_name in SKIP_SHEETS:
        continue
    
    print(f"\nProcessing sheet: {sheet_name}")
    ws = wb[sheet_name]
    grade_key = sheet_name.replace(' ', '').lower()
    grade_file = f"{grade_key}.json"
    publications = []
    
    headers = [str(cell.value).strip() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"Found columns: {headers}")

    row_count = 0
    valid_lessons = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1
        data = dict(zip(headers, row))
        data = {k: (v.strip() if isinstance(v, str) else v) for k, v in data.items()}
        
        lesson_id = str(data.get('lesson_id', '')).strip()
        title = str(data.get('title ') or data.get('lesson_name') or '').strip()
        asset = str(data.get('Asset Link', '')).strip()
        cocos_lesson_code = str(data.get('cocos_lesson_code', '')).strip() or data.get('id') or 'default'

        if not lesson_id or lesson_id.lower() == 'nan':
            print(f"Skipping row {row_count}: Missing lesson_id")
            continue
            
        if not title or title.lower() == 'nan':
            title = f"Lesson {lesson_id}"
            print(f"Row {row_count}: Using default title: {title}")
            
        if not asset or asset.lower() == 'nan':
            print(f"Skipping row {row_count}: Missing asset link")
            continue

        valid_lessons += 1
        print(f"Processing lesson {lesson_id}: {title}")

        lesson_filename = f"{lesson_id}.json"

        image_base = urljoin(BASE_URL, f"images/icons/{cocos_lesson_code}")
        images = [{
            "href": f"{image_base}.jpg",
            "type": "image/jpeg",
            "height": 1400,
            "width": 800
        }]

        publication = {
            'metadata': {
                '@type': 'http://schema.org/Game',
                'title': title,
                'author': str(data.get('author', '')).strip() or 'Chimple',
                'identifier': str(data.get('identifier_url', '')).strip() or urljoin(BASE_URL, f"id/{lesson_id}"),
                'language': str(data.get('language_id', '')).strip() or 'en',
                'modified': (data.get('modified').isoformat() if isinstance(data.get('modified'), datetime)
                             else datetime.now().isoformat())
            },
            'links': [
                {'rel': 'self', 'href': urljoin(BASE_URL, lesson_filename), 'type': PUB_TYPE}
            ],
            'images': images
        }
        publications.append(publication)

        lesson_manifest = {
            'metadata': publication['metadata'],
            'links': publication['links'],
            'images': images,
            'resources': [{'href': asset, 'type': 'application/zip'}]
        }

        with open(os.path.join(LESSON_DIR, lesson_filename), 'w', encoding='utf-8') as lf:
            json.dump(lesson_manifest, lf, indent=2)
        print(f"Generated lesson manifest: {lesson_filename}")

    grade_json = {
        'metadata': {'title': f"{sheet_name}"},
        'links': [{'rel': 'self', 'href': urljoin(BASE_URL, grade_file), 'type': TYPE_OPDS}],
        'publications': publications
    }
    with open(os.path.join(GRADE_DIR, grade_file), 'w', encoding='utf-8') as gf:
        json.dump(grade_json, gf, indent=2)
    print(f"Generated {grade_file} with {len(publications)} valid lessons (skipped {row_count - valid_lessons} rows)")

print("\nAll files generated successfully!")
